from decimal import Decimal
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from catalog.models import Brand, Category, MediaImage, Product


class PromotionZoneTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = get_user_model().objects.create_superuser(
            username='admin',
            password='pwd',
        )
        self.major_category = Category.objects.create(
            name='家电',
            level=Category.LEVEL_MAJOR,
        )
        self.category = Category.objects.create(
            name='厨房电器',
            level=Category.LEVEL_MINOR,
            parent=self.major_category,
        )
        self.brand = Brand.objects.create(name='测试品牌')

    def create_product(self, name, **overrides):
        data = {
            'name': name,
            'category': self.category,
            'brand': self.brand,
            'price': Decimal('99.00'),
            'stock': 10,
            'is_active': True,
        }
        data.update(overrides)
        return Product.objects.create(**data)

    def create_media(self, filename='promotion.jpg'):
        return MediaImage.objects.create(
            file=f'images/{filename}',
            original_name=filename,
            content_type='image/jpeg',
            size=128,
        )

    def get_results(self, response):
        data = response.json()
        if isinstance(data, dict):
            return data.get('results', data.get('data', []))
        return data

    def test_product_list_filters_promotion_zone_products(self):
        promotion_product = self.create_product(
            '活动商品',
            show_in_promotion_zone=True,
        )
        self.create_product('普通商品', show_in_promotion_zone=False)

        response = self.client.get(
            '/api/catalog/products/',
            {'show_in_promotion_zone': 'true', 'page_size': 10},
        )

        self.assertEqual(response.status_code, 200, response.content)
        data = response.json()
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['results'][0]['id'], promotion_product.id)
        self.assertTrue(data['results'][0]['show_in_promotion_zone'])

    def test_product_export_filters_and_includes_promotion_zone_column(self):
        self.client.force_authenticate(self.admin)
        promotion_product = self.create_product(
            '活动导出商品',
            show_in_promotion_zone=True,
        )
        self.create_product('普通导出商品', show_in_promotion_zone=False)

        response = self.client.get(
            '/api/catalog/products/export/',
            {'show_in_promotion_zone': 'true'},
        )

        self.assertEqual(response.status_code, 200, response.content)
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[2]]
        self.assertIn('优惠专区', headers)
        product_id_col = headers.index('商品ID') + 1
        promotion_col = headers.index('优惠专区') + 1
        data_rows = list(sheet.iter_rows(min_row=3))
        self.assertEqual(len(data_rows), 1)
        self.assertEqual(data_rows[0][product_id_col - 1].value, promotion_product.id)
        self.assertEqual(data_rows[0][promotion_col - 1].value, '是')

    def test_home_banner_accepts_promotion_position(self):
        self.client.force_authenticate(self.admin)
        media = self.create_media('promotion-banner.jpg')

        response = self.client.post(
            '/api/catalog/home-banners/',
            {
                'image': media.id,
                'title': '优惠专区轮播',
                'position': 'promotion',
                'order': 1,
                'is_active': True,
            },
            format='json',
        )

        self.assertEqual(response.status_code, 201, response.content)
        self.assertEqual(response.json()['position'], 'promotion')

        public_client = APIClient()
        list_response = public_client.get(
            '/api/catalog/home-banners/',
            {'position': 'promotion'},
        )
        self.assertEqual(list_response.status_code, 200, list_response.content)
        results = self.get_results(list_response)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['position'], 'promotion')

    def test_special_zone_cover_accepts_promotion_type(self):
        self.client.force_authenticate(self.admin)
        media = self.create_media('promotion-cover.jpg')

        response = self.client.post(
            '/api/catalog/special-zone-covers/',
            {
                'image': media.id,
                'type': 'promotion',
                'is_active': True,
            },
            format='json',
        )

        self.assertEqual(response.status_code, 201, response.content)
        self.assertEqual(response.json()['type'], 'promotion')

        public_client = APIClient()
        list_response = public_client.get(
            '/api/catalog/special-zone-covers/',
            {'type': 'promotion'},
        )
        self.assertEqual(list_response.status_code, 200, list_response.content)
        results = self.get_results(list_response)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['type'], 'promotion')
