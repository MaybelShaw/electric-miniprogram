from __future__ import annotations

from typing import Iterable, Sequence

from django.http import HttpResponse


def build_excel_response(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    title: str | None = None,
) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    row_index = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        row_index += 1

    max_lens = [len(str(h or "")) for h in headers]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_index, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    data_row = row_index + 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            cell_value = "" if value is None else value
            ws.cell(row=data_row, column=col_idx, value=cell_value)
            max_lens[col_idx - 1] = max(max_lens[col_idx - 1], len(str(cell_value)))
        data_row += 1

    for col_idx, max_len in enumerate(max_lens, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
