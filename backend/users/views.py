from django.shortcuts import render
import requests
import uuid
from rest_framework import viewsets, status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action, api_view, permission_classes, throttle_classes
from django.conf import settings
from django.utils import timezone
from .models import User, Address, CompanyInfo, CreditAccount, AccountStatement, AccountTransaction, Notification
from orders.models import Order
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth, TruncYear
from .serializers import (
    UserSerializer,
    UserProfileSerializer,
    AddressSerializer,
    CompanyInfoSerializer,
    CreditAccountSerializer,
    AccountStatementSerializer,
    AccountStatementDetailSerializer,
    AccountTransactionSerializer,
    NotificationSerializer,
)
from django.db.models import Q
from common.permissions import IsOwnerOrAdmin, IsAdmin
from common.throttles import LoginRateThrottle
from common.address_parser import address_parser
from common.utils import to_bool
from common.excel import build_excel_response
from common.pagination import SmallResultsSetPagination
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes as OT


# Create your views here.
@extend_schema(tags=['Authentication'])
class WeChatLoginView(APIView):
    """
    微信小程序登录接口
    
    功能：
    - 接收微信小程序的code
    - 调用微信API验证code并获取openid
    - 创建或获取用户
    - 返回JWT token
    
    登录模式：
    - 配置了微信凭证：调用真实微信API（开发和生产环境推荐）
    - 未配置微信凭证：使用模拟登录（仅用于本地测试）
    
    特殊功能：
    - code以'admin'开头时自动授予管理员权限（仅开发环境）
    """
    # 登录接口允许匿名访问，并关闭 JWT 认证，避免因携带无效令牌导致 401
    permission_classes = [permissions.AllowAny]
    authentication_classes: list = []
    throttle_classes = [LoginRateThrottle]
    
    @extend_schema(
        operation_id='wechat_login',
        description='WeChat mini program login. Exchange WeChat code for JWT token.',
    )
    def post(self, request):
        import logging
        logger = logging.getLogger(__name__)
        
        code = request.data.get("code")
        if not code:
            logger.warning('微信登录失败: 缺少code参数')
            return Response(
                {"error": "Code is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        # 获取微信配置
        appid = settings.WECHAT_APPID
        secret = settings.WECHAT_SECRET
        
        # 检查是否配置了微信凭证
        if not appid or not secret:
            if settings.DEBUG:
                logger.warning('微信凭证未配置，使用模拟登录模式')
                logger.info(f'模拟登录: code={code}')
                openid = code
                session_key = None
            else:
                logger.error('微信凭证未配置，生产环境禁止模拟登录')
                return Response(
                    {"error": "WeChat credentials are not configured"},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE
                )
        else:
            # 调用微信API（开发和生产环境都使用真实API）
            url = f"https://api.weixin.qq.com/sns/jscode2session?appid={appid}&secret={secret}&js_code={code}&grant_type=authorization_code"
            
            try:
                logger.info(f'调用微信API: code={code[:10]}...')
                response = requests.get(url, timeout=10)
                data = response.json()
                
                logger.info(f'微信API响应: keys={list(data.keys())}')
                
                if "errcode" in data:
                    error_code = data.get("errcode")
                    error_msg = data.get("errmsg", "Unknown error")
                    logger.error(f'微信API错误: errcode={error_code}, errmsg={error_msg}')
                    return Response(
                        {
                            "error": "WeChat API error",
                            "errcode": error_code,
                            "errmsg": error_msg
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                openid = data.get("openid")
                session_key = data.get("session_key")
                
                if not openid:
                    logger.error(f'微信API返回数据异常: 缺少openid, data={data}')
                    return Response(
                        {"error": "Invalid WeChat response"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                
                logger.info(f'微信登录成功: openid={openid[:10]}...')
                
            except requests.RequestException as e:
                logger.error(f'调用微信API失败: {str(e)}')
                return Response(
                    {"error": "Failed to connect to WeChat API"},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE
                )
            except Exception as e:
                logger.error(f'微信登录异常: {str(e)}')
                return Response(
                    {"error": "Internal server error"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        from .services import (
            get_or_create_wechat_user,
            grant_admin_on_debug_code,
            update_last_login,
            create_tokens_for_user,
        )

        user, created = get_or_create_wechat_user(openid)
        
        if created:
            logger.info(f'创建新用户: user_id={user.id}, openid={openid[:10]}...')
        else:
            logger.info(f'用户登录: user_id={user.id}, openid={openid[:10]}...')

        # 开发环境支持管理员快捷登录：code 以 'admin' 开头时授予管理员权限
        try:
            granted = grant_admin_on_debug_code(str(code).strip().lower(), user, settings.DEBUG)
            if granted:
                logger.info(f'授予管理员权限: user_id={user.id}')
        except Exception as e:
            logger.error(f'授予管理员权限失败: {str(e)}')

        update_last_login(user)

        refresh, access = create_tokens_for_user(user)

        logger.info(f'登录成功: user_id={user.id}, username={user.username}')

        return Response(
            {
                "access": access,
                "refresh": refresh,
                "user": UserSerializer(user).data,
            }
        )


from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

@extend_schema(tags=['Authentication'])
@method_decorator(csrf_exempt, name='dispatch')
class PasswordLoginView(APIView):
    """管理员及客服用户名+密码登录。
    - 允许匿名访问
    - 验证用户名与密码
    - 非管理员或客服用户返回 403
    - 开发环境支持快捷创建管理员（若用户名不存在）
    """
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    throttle_classes = [LoginRateThrottle]

    @extend_schema(
        operation_id='password_login',
        description='Admin and Support password login. Authenticate with username and password to get JWT token.',
    )
    def post(self, request):
        from .services import (
            bootstrap_or_init_admin,
            ensure_user_password,
            update_last_login,
            create_tokens_for_user,
        )
        username = request.data.get("username")
        password = request.data.get("password")
        if not username or not password:
            return Response({"error": "用户名与密码必填"}, status=status.HTTP_400_BAD_REQUEST)

        admin_exists = User.objects.filter(is_staff=True).exists()
        user = User.objects.filter(username=username).order_by('-date_joined').first()

        if user is None:
            created_or_initialized = bootstrap_or_init_admin(username, password)
            if created_or_initialized:
                update_last_login(created_or_initialized)
                refresh, access = create_tokens_for_user(created_or_initialized)
                return Response(
                    {
                        "access": access,
                        "refresh": refresh,
                        "user": UserSerializer(created_or_initialized).data,
                    }
                )
            return Response({"error": "用户名或密码错误"}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            # 用户存在但无可用密码：允许使用当前提交的密码进行首次初始化（不自动提升为管理员，除非系统无管理员）
            ensure_user_password(user, password)

        # 校验密码
        if not user.check_password(password):
            return Response({"error": "用户名或密码错误"}, status=status.HTTP_401_UNAUTHORIZED)

        # 权限校验：仅在系统无管理员的情况下提升为管理员以完成首次引导
        if not user.is_staff:
            if not admin_exists:
                user.is_staff = True
                user.is_superuser = True
                user.role = 'admin'
                user.save()
            elif getattr(user, 'role', '') != 'support':
                return Response({"error": "无管理员权限"}, status=status.HTTP_403_FORBIDDEN)
        
        # 确保管理员用户的 role 字段正确
        if user.is_staff and user.role != 'admin':
            user.role = 'admin'
            user.save(update_fields=['role'])

        update_last_login(user)

        refresh, access = create_tokens_for_user(user)

        return Response(
            {
                "access": access,
                "refresh": refresh,
                "user": UserSerializer(user).data,
            }
        )

@api_view(['GET','PATCH'])
@permission_classes([IsAuthenticated])
def user_profile(request):
    user = request.user
    if request.method == 'GET':
        serializer = UserProfileSerializer(user)
        return Response(serializer.data)
    elif request.method == 'PATCH':
        serializer = UserProfileSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_statistics(request):
    """
    Get user statistics including orders.
    
    Returns:
    - orders_count: Total number of orders
    - completed_orders_count: Number of completed orders
    - pending_orders_count: Number of pending orders
    - total_spent: Total amount spent on completed orders
    """
    from django.core.cache import cache
    from django.db.models import Sum
    
    user = request.user
    cache_key = f'user_stats_{user.id}'
    
    stats = cache.get(cache_key)
    if stats is None:
        orders_qs = user.orders.all()
        
        stats = {
            'orders_count': orders_qs.count(),
            'completed_orders_count': orders_qs.filter(status='completed').count(),
            'pending_orders_count': orders_qs.filter(status='pending').count(),
            'total_spent': float(
                orders_qs.filter(status='completed').aggregate(
                    total=Sum('total_amount')
                )['total'] or 0
            ),
        }
        
        cache.set(cache_key, stats, 300)  # Cache for 5 minutes
    
    return Response(stats)

@extend_schema(tags=['Addresses'])
class AddressViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing user addresses.
    
    Permissions:
    - IsAuthenticated: Users can only manage their own addresses
    
    Note: list() method returns a simple array instead of paginated response
    """
    queryset = Address.objects.all()
    serializer_class = AddressSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None  # 禁用分页

    def get_queryset(self):
        return Address.objects.filter(user=self.request.user).order_by('-is_default', '-id')

    def list(self, request, *args, **kwargs):
        """返回地址列表（数组格式，不分页）"""
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    def set_default(self, request, pk=None):
        address = self.get_object()
        Address.objects.filter(user=request.user, is_default=True).update(
            is_default=False
        )
        address.is_default = True
        address.save()
        return Response({"status": "默认地址已设置"})
    
    @extend_schema(
        operation_id='parse_address',
        description='Parse a complete address text into province, city, district and detail',
        request={
            'application/json': {
                'type': 'object',
                'properties': {
                    'address': {
                        'type': 'string',
                        'description': '完整的地址文本',
                        'example': '北京市朝阳区建国路88号SOHO现代城A座1001室'
                    }
                },
                'required': ['address']
            }
        },
        responses={
            200: {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'},
                    'data': {
                        'type': 'object',
                        'properties': {
                            'province': {'type': 'string', 'nullable': True},
                            'city': {'type': 'string', 'nullable': True},
                            'district': {'type': 'string', 'nullable': True},
                            'detail': {'type': 'string', 'nullable': True}
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=["post"], url_path="parse")
    def parse_address(self, request):
        """
        解析完整地址文本
        
        用户输入完整地址后，后端自动识别并拆分成省、市、区、详细地址
        """
        address_text = request.data.get('address', '').strip()
        
        if not address_text:
            return Response({
                'success': False,
                'message': '地址不能为空',
                'data': None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 使用地址解析器解析地址
        result = address_parser.parse_address(address_text)
        
        return Response({
            'success': result['success'],
            'message': result['message'],
            'data': {
                'province': result['province'],
                'city': result['city'],
                'district': result['district'],
                'detail': result['detail']
            }
        })

    @extend_schema(
        operation_id='addresses_destroy',
        description='删除地址：如被其他数据引用时，返回提示信息',
    )
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            from django.db.models.deletion import ProtectedError
            if isinstance(e, ProtectedError):
                return Response(
                    {
                        'error': '无法删除地址',
                        'message': '该地址被其他数据引用，无法删除',
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            raise


@extend_schema(tags=['Notifications'])
class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    用户消息中心
    - 支持按类型/状态筛选
    - 支持标记单条或全部已读
    - 提供订阅模板配置
    """
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = SmallResultsSetPagination

    def get_queryset(self):
        qs = Notification.objects.filter(user=self.request.user).order_by('-created_at')

        ntype = self.request.query_params.get('type')
        if ntype:
            qs = qs.filter(type=ntype)

        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        read_flag = self.request.query_params.get('read')
        if read_flag is not None:
            val = str(read_flag).lower()
            if val in {'1', 'true', 'yes', 'read'}:
                qs = qs.filter(read_at__isnull=False)
            elif val in {'0', 'false', 'no', 'unread'}:
                qs = qs.filter(read_at__isnull=True)

        return qs

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notif = self.get_object()
        notif.mark_read()
        return Response(self.get_serializer(notif).data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        updated = Notification.objects.filter(
            user=request.user,
            read_at__isnull=True
        ).update(read_at=timezone.now())
        return Response({'marked': updated})

    @action(detail=False, methods=['get'])
    def stats(self, request):
        qs = Notification.objects.filter(user=request.user)
        return Response({
            'unread_count': qs.filter(read_at__isnull=True).count(),
            'pending_count': qs.filter(status='pending').count(),
            'total': qs.count(),
        })

    @action(detail=False, methods=['get'])
    def subscribe_templates(self, request):
        """
        返回后端配置的订阅消息模板ID，便于前端申请授权。
        """
        templates = getattr(settings, 'WECHAT_SUBSCRIBE_TEMPLATES', {}) or {}
        default_page = getattr(settings, 'WECHAT_SUBSCRIBE_DEFAULT_PAGE', '')

        items = []
        if isinstance(templates, dict):
            for scene, tmpl in templates.items():
                if isinstance(tmpl, dict):
                    items.append({
                        'scene': scene,
                        'template_id': tmpl.get('template_id') or tmpl.get('id') or '',
                        'page': tmpl.get('page') or default_page,
                    })
                else:
                    items.append({
                        'scene': scene,
                        'template_id': str(tmpl),
                        'page': default_page,
                    })

        return Response({
            'templates': items,
            'default_page': default_page,
        })


@extend_schema(tags=['Users'])
class AdminUserViewSet(viewsets.ModelViewSet):
    """
    ViewSet for administrator user management.
    
    Permissions:
    - IsAuthenticated: User must be logged in
    - IsAdmin: User must be an administrator
    
    Provides CRUD operations and admin privilege management.
    """
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        # 统一搜索：在用户名或 OpenID 上模糊匹配
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(username__icontains=search) | Q(openid__icontains=search))
        # 新增：支持按手机号模糊查询
        phone = self.request.query_params.get('phone')
        if phone:
            qs = qs.filter(phone__icontains=phone)
        # 公司名筛选
        company_name = self.request.query_params.get('company_name')
        if company_name:
            qs = qs.filter(
                company_info__status='approved',
                company_info__company_name__icontains=company_name,
            )
        # 管理员筛选：支持 true/false/1/0/布尔
        is_staff = self.request.query_params.get('is_staff')
        if is_staff is not None:
            parsed = to_bool(is_staff)
            if parsed is not None:
                try:
                    qs = qs.filter(is_staff=parsed)
                except Exception:
                    pass
        # 角色筛选：支持按role字段筛选
        role = self.request.query_params.get('role')
        if role:
            try:
                qs = qs.filter(role=role)
            except Exception:
                pass
        return qs

    @action(detail=False, methods=['get'])
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())
        headers = [
            '用户名',
            '手机号',
            '邮箱',
            '角色',
            '管理员',
            '注册时间',
            '最后登录',
        ]
        rows = []
        for user in qs:
            rows.append([
                user.username,
                user.phone,
                user.email,
                user.get_role_display() if hasattr(user, 'get_role_display') else user.role,
                '是' if user.is_staff else '否',
                user.date_joined,
                user.last_login_at,
            ])
        filename = f"users_export_{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return build_excel_response(filename, headers, rows, title="用户导出")

    def create(self, request, *args, **kwargs):
        # 允许管理员创建用户：当未提供 openid 时自动生成，若提供密码则安全哈希
        data = request.data.copy()
        if not data.get('openid'):
            data['openid'] = f"manual:{uuid.uuid4().hex[:24]}"
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        instance = serializer.save()
        raw_password = self.request.data.get('password')
        if raw_password:
            try:
                instance.set_password(raw_password)
                instance.save()
            except Exception:
                # 若设置密码失败，不阻断创建；密码保持未初始化，由后续流程设置
                pass

    @action(detail=True, methods=['post'])
    def set_admin(self, request, pk=None):
        user = self.get_object()
        user.is_staff = True
        user.is_superuser = True
        user.role = 'admin'
        user.save()
        return Response(UserSerializer(user).data)

    @action(detail=True, methods=['post'])
    def unset_admin(self, request, pk=None):
        user = self.get_object()
        user.is_staff = False
        user.is_superuser = False
        # When removing admin privileges, reset to individual unless they are a dealer
        if user.role == 'admin':
            user.role = 'dealer' if hasattr(user, 'company_info') and user.company_info.status == 'approved' else 'individual'
        user.save()
        return Response(UserSerializer(user).data)

    @action(detail=True, methods=['post'])
    def force_delete(self, request, pk=None):
        user = self.get_object()
        user_id = user.id
        from django.db import transaction
        from django.db.models.deletion import ProtectedError
        from django.core.files.storage import default_storage
        from orders.models import (
            Order,
            OrderStatusHistory,
            Payment,
            Refund,
            Invoice,
            ReturnRequest,
            Cart,
            CartItem,
            DiscountTarget,
        )
        from support.models import SupportConversation, SupportMessage
        from catalog.models import SearchLog, InventoryLog

        files_to_delete = set()

        def queue_file(file_field):
            name = getattr(file_field, 'name', '')
            if name:
                files_to_delete.add(name)

        try:
            with transaction.atomic():
                for invoice in Invoice.objects.filter(user=user).only('id', 'file'):
                    queue_file(invoice.file)
                for message in SupportMessage.objects.filter(
                    Q(conversation__user=user) | Q(sender=user)
                ).only('id', 'attachment'):
                    queue_file(message.attachment)

                # 清理作为操作人的关联，避免 PROTECT 阻塞
                Refund.objects.filter(operator=user).update(operator=None)
                ReturnRequest.objects.filter(processed_by=user).update(processed_by=None)
                OrderStatusHistory.objects.filter(operator=user).update(operator=None)

                # 客服与行为日志
                SupportMessage.objects.filter(sender=user).delete()
                SupportConversation.objects.filter(user=user).delete()
                SearchLog.objects.filter(user=user).delete()
                InventoryLog.objects.filter(created_by=user).delete()

                # 用户基础数据
                Notification.objects.filter(user=user).delete()
                Address.objects.filter(user=user).delete()
                DiscountTarget.objects.filter(user=user).delete()

                # 购物车
                CartItem.objects.filter(cart__user=user).delete()
                Cart.objects.filter(user=user).delete()

                # 订单与财务数据
                Refund.objects.filter(order__user=user).delete()
                Payment.objects.filter(order__user=user).delete()
                Invoice.objects.filter(user=user).delete()
                ReturnRequest.objects.filter(order__user=user).delete()
                OrderStatusHistory.objects.filter(order__user=user).delete()
                Order.objects.filter(user=user).delete()

                credit_account = getattr(user, 'credit_account', None)
                if credit_account:
                    AccountTransaction.objects.filter(credit_account=credit_account).delete()
                    AccountStatement.objects.filter(credit_account=credit_account).delete()
                    credit_account.delete()

                company_info = getattr(user, 'company_info', None)
                if company_info:
                    company_info.delete()

                user.delete()

                if files_to_delete:
                    files_snapshot = list(files_to_delete)

                    def _cleanup_files():
                        for path in files_snapshot:
                            try:
                                default_storage.delete(path)
                            except Exception:
                                pass

                    transaction.on_commit(_cleanup_files)
        except ProtectedError:
            return Response(
                {"error": "强制删除失败", "message": "仍存在关联数据，无法删除"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as exc:
            return Response(
                {"error": "强制删除失败", "message": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({"message": "用户及其关联数据已强制删除", "user_id": user_id})

    @action(detail=True, methods=['get'])
    def transaction_stats(self, request, pk=None):
        try:
            period = request.query_params.get('period', 'month')
            include_paid = request.query_params.get('include_paid', 'false')
            include_paid = str(include_paid).lower() in ['true', '1']
            year = request.query_params.get('year')
            user_id = int(pk)
        except Exception:
            return Response({'detail': 'invalid parameters'}, status=status.HTTP_400_BAD_REQUEST)

        qs = Order.objects.filter(user_id=user_id)
        statuses = ['completed']
        if include_paid:
            statuses.append('paid')
        qs = qs.filter(status__in=statuses)
        if year:
            qs = qs.filter(created_at__year=year)

        if period == 'year':
            data = list(
                qs.annotate(p=TruncYear('created_at'))
                .values('p')
                .annotate(orders=Count('id'), amount=Sum('total_amount'))
                .order_by('p')
            )
            result = [
                {
                    'period': d['p'].year if hasattr(d['p'], 'year') else str(d['p']),
                    'orders': d['orders'] or 0,
                    'amount': float(d['amount'] or 0)
                } for d in data
            ]
            return Response(result)

        data = list(
            qs.annotate(p=TruncMonth('created_at'))
            .values('p')
            .annotate(orders=Count('id'), amount=Sum('total_amount'))
            .order_by('p')
        )
        result = [
            {
                'period': d['p'].strftime('%Y-%m') if hasattr(d['p'], 'strftime') else str(d['p']),
                'orders': d['orders'] or 0,
                'amount': float(d['amount'] or 0)
            } for d in data
        ]
        return Response(result)

    @action(detail=False, methods=['get'])
    def customers_transaction_stats(self, request):
        if not request.user.is_staff:
            return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

        period = request.query_params.get('period', 'month')
        include_paid = request.query_params.get('include_paid', 'false')
        include_paid = str(include_paid).lower() in ['true', '1']
        year = request.query_params.get('year')
        role = request.query_params.get('role')

        qs = Order.objects.all()
        statuses = ['completed']
        if include_paid:
            statuses.append('paid')
        qs = qs.filter(status__in=statuses)
        if year:
            qs = qs.filter(created_at__year=year)
        if role:
            qs = qs.filter(user__role=role)

        if period == 'year':
            data = list(
                qs.annotate(p=TruncYear('created_at'))
                .values('user_id', 'user__username', 'p')
                .annotate(orders=Count('id'), amount=Sum('total_amount'))
                .order_by('user_id', 'p')
            )
            result = [
                {
                    'user_id': d['user_id'],
                    'username': d['user__username'],
                    'period': d['p'].year if hasattr(d['p'], 'year') else str(d['p']),
                    'orders': d['orders'] or 0,
                    'amount': float(d['amount'] or 0)
                } for d in data
            ]
            return Response(result)

        data = list(
            qs.annotate(p=TruncMonth('created_at'))
            .values('user_id', 'user__username', 'p')
            .annotate(orders=Count('id'), amount=Sum('total_amount'))
            .order_by('user_id', 'p')
        )
        result = [
            {
                'user_id': d['user_id'],
                'username': d['user__username'],
                'period': d['p'].strftime('%Y-%m') if hasattr(d['p'], 'strftime') else str(d['p']),
                'orders': d['orders'] or 0,
                'amount': float(d['amount'] or 0)
            } for d in data
        ]
        return Response(result)

    @action(detail=True, methods=['get'])
    def export_transaction_stats(self, request, pk=None):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        try:
            period = request.query_params.get('period', 'month')
            include_paid = request.query_params.get('include_paid', 'false')
            include_paid = str(include_paid).lower() in ['true', '1']
            year = request.query_params.get('year')
            user = self.get_object()
        except Exception:
            return Response({'detail': 'invalid parameters'}, status=status.HTTP_400_BAD_REQUEST)

        qs = Order.objects.filter(user_id=user.id)
        statuses = ['completed']
        if include_paid:
            statuses.append('paid')
        qs = qs.filter(status__in=statuses)
        if year:
            qs = qs.filter(created_at__year=year)

        if period == 'year':
            data = list(
                qs.annotate(p=TruncYear('created_at'))
                .values('p')
                .annotate(orders=Count('id'), amount=Sum('total_amount'))
                .order_by('p')
            )
            rows = [
                [
                    (d['p'].year if hasattr(d['p'], 'year') else str(d['p'])),
                    d['orders'] or 0,
                    float(d['amount'] or 0)
                ] for d in data
            ]
            header_text = '年份'
        else:
            data = list(
                qs.annotate(p=TruncMonth('created_at'))
                .values('p')
                .annotate(orders=Count('id'), amount=Sum('total_amount'))
                .order_by('p')
            )
            rows = [
                [
                    (d['p'].strftime('%Y-%m') if hasattr(d['p'], 'strftime') else str(d['p'])),
                    d['orders'] or 0,
                    float(d['amount'] or 0)
                ] for d in data
            ]
            header_text = '月份'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '交易统计'
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        ws.merge_cells('A1:C1')
        ws['A1'] = f"客户交易统计 - {user.username}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A3'] = '统计维度'
        ws['B3'] = '按年' if period == 'year' else '按月'
        ws['A4'] = '筛选年份'
        ws['B4'] = year or ''
        ws['A5'] = '包含已支付未完成'
        ws['B5'] = '是' if include_paid else '否'
        headers = [header_text, '订单数', '交易金额']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        start_row = 8
        for i, row in enumerate(rows, start=start_row):
            ws.cell(row=i, column=1, value=row[0])
            ws.cell(row=i, column=2, value=row[1])
            ws.cell(row=i, column=3, value=row[2])
        if rows:
            total_orders = sum(r[1] for r in rows)
            total_amount = sum(r[2] for r in rows)
            ws.cell(row=start_row + len(rows), column=1, value='总计')
            ws.cell(row=start_row + len(rows), column=2, value=total_orders)
            ws.cell(row=start_row + len(rows), column=3, value=float(total_amount))
        for idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"user_{user.id}_transaction_stats_{period}"
        if year:
            filename += f"_{year}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_customers_transaction_stats(self, request):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        if not request.user.is_staff:
            return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)
        period = request.query_params.get('period', 'month')
        include_paid = request.query_params.get('include_paid', 'false')
        include_paid = str(include_paid).lower() in ['true', '1']
        year = request.query_params.get('year')
        role = request.query_params.get('role')
        qs = Order.objects.all()
        statuses = ['completed']
        if include_paid:
            statuses.append('paid')
        qs = qs.filter(status__in=statuses)
        if year:
            qs = qs.filter(created_at__year=year)
        if role:
            qs = qs.filter(user__role=role)
        if period == 'year':
            data = list(
                qs.annotate(p=TruncYear('created_at'))
                .values('user_id', 'user__username', 'p')
                .annotate(orders=Count('id'), amount=Sum('total_amount'))
                .order_by('user_id', 'p')
            )
            rows = [
                [
                    d['user__username'],
                    (d['p'].year if hasattr(d['p'], 'year') else str(d['p'])),
                    d['orders'] or 0,
                    float(d['amount'] or 0)
                ] for d in data
            ]
            header_text = '年份'
        else:
            data = list(
                qs.annotate(p=TruncMonth('created_at'))
                .values('user_id', 'user__username', 'p')
                .annotate(orders=Count('id'), amount=Sum('total_amount'))
                .order_by('user_id', 'p')
            )
            rows = [
                [
                    d['user__username'],
                    (d['p'].strftime('%Y-%m') if hasattr(d['p'], 'strftime') else str(d['p'])),
                    d['orders'] or 0,
                    float(d['amount'] or 0)
                ] for d in data
            ]
            header_text = '月份'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '客户交易统计'
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        ws.merge_cells('A1:D1')
        ws['A1'] = '客户交易统计汇总'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A3'] = '统计维度'
        ws['B3'] = '按年' if period == 'year' else '按月'
        ws['A4'] = '筛选年份'
        ws['B4'] = year or ''
        ws['A5'] = '包含已支付未完成'
        ws['B5'] = '是' if include_paid else '否'
        headers = ['用户名', header_text, '订单数', '交易金额']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        start_row = 8
        for i, row in enumerate(rows, start=start_row):
            ws.cell(row=i, column=1, value=row[0])
            ws.cell(row=i, column=2, value=row[1])
            ws.cell(row=i, column=3, value=row[2])
            ws.cell(row=i, column=4, value=row[3])
        if rows:
            total_orders = sum(r[2] for r in rows)
            total_amount = sum(r[3] for r in rows)
            ws.cell(row=start_row + len(rows), column=1, value='总计')
            ws.cell(row=start_row + len(rows), column=3, value=total_orders)
            ws.cell(row=start_row + len(rows), column=4, value=float(total_amount))
        for idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"customers_transaction_stats_{period}"
        if year:
            filename += f"_{year}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    @extend_schema(
        operation_id='users_destroy',
        description='删除用户：存在关联数据时，阻止删除并返回提示信息',
    )
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        addresses_count = instance.addresses.count()
        orders_count = instance.orders.count()
        cart_count = instance.cart.count()
        discount_targets_count = instance.discount_targets.count()
        company_info_exists = 1 if hasattr(instance, 'company_info') else 0
        credit_account_exists = 1 if hasattr(instance, 'credit_account') else 0
        statements_count = instance.credit_account.statements.count() if credit_account_exists else 0
        transactions_count = instance.credit_account.transactions.count() if credit_account_exists else 0
        total_refs = addresses_count + orders_count + cart_count + discount_targets_count + company_info_exists + credit_account_exists + statements_count + transactions_count
        if total_refs > 0:
            return Response(
                {
                    'error': '无法删除用户',
                    'message': (
                        f"该用户存在 {addresses_count} 个地址、{orders_count} 个订单、{cart_count} 个购物车、{discount_targets_count} 条折扣关联、{company_info_exists} 条公司信息、{credit_account_exists} 个信用账户、{statements_count} 条对账单、{transactions_count} 条账务交易，无法删除"
                    ).strip(),
                    'addresses_count': addresses_count,
                    'orders_count': orders_count,
                    'cart_count': cart_count,
                    'discount_targets_count': discount_targets_count,
                    'company_info_exists': company_info_exists,
                    'credit_account_exists': credit_account_exists,
                    'statements_count': statements_count,
                    'transactions_count': transactions_count,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            from django.db.models.deletion import ProtectedError
            if isinstance(e, ProtectedError):
                return Response(
                    {
                        'error': '无法删除用户',
                        'message': '该用户被其他数据引用，无法删除',
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            raise


@extend_schema(tags=['Company Info'])
class CompanyInfoViewSet(viewsets.ModelViewSet):
    """
    ViewSet for company information management.
    
    User endpoints:
    - POST: Submit company info for dealer certification
    - GET: View own company info
    - PATCH: Update company info (only if rejected or withdrawn)
    
    Admin endpoints:
    - GET list: View all company info submissions
    - approve: Approve company info and upgrade user to dealer
    - reject: Reject company info submission
    - withdraw: Withdraw pending submission
    """
    queryset = CompanyInfo.objects.all().order_by('-created_at')
    serializer_class = CompanyInfoSerializer
    
    def get_permissions(self):
        """
        Override to allow different permissions for different actions
        """
        if self.action in ['approve', 'reject']:
            # Admin-only actions
            return [IsAuthenticated(), IsAdmin()]
        # Regular authenticated users can create/view/update their own
        return [IsAuthenticated()]
    
    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return CompanyInfo.objects.none()
        
        if user.is_staff:
            # Admin can see all
            qs = super().get_queryset()
            # Filter by status
            status_filter = self.request.query_params.get('status')
            if status_filter:
                qs = qs.filter(status=status_filter)
            return qs
        else:
            # Regular users can only see their own
            return CompanyInfo.objects.filter(user=user)
    
    def list(self, request, *args, **kwargs):
        """List company info - admin sees all, users see their own"""
        if not request.user.is_staff:
            # For regular users, return their company info or empty
            try:
                company_info = CompanyInfo.objects.get(user=request.user)
                serializer = self.get_serializer(company_info)
                return Response(serializer.data)
            except CompanyInfo.DoesNotExist:
                return Response(None, status=status.HTTP_404_NOT_FOUND)
        
        # Admin gets paginated list
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        """Submit company info for certification"""
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f'创建公司信息请求: user={request.user}, is_authenticated={request.user.is_authenticated}')
        
        # Check if user already has company info
        if hasattr(request.user, 'company_info'):
            instance = request.user.company_info
            logger.warning(f'用户已有公司信息: user_id={request.user.id}, status={instance.status}')
            if instance.status == 'approved':
                return Response(
                    {"error": "已审核通过的信息不可重复提交"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if instance.status == 'pending':
                return Response(
                    {"error": "审核中不可修改，请先撤回"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Treat create as resubmission when status is rejected/withdrawn
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            update_fields = ['updated_at']
            if instance.status != 'pending':
                instance.status = 'pending'
                instance.approved_at = None
                update_fields.extend(['status', 'approved_at'])
            if instance.reject_reason:
                instance.reject_reason = ''
                update_fields.append('reject_reason')
            instance.save(update_fields=update_fields)
            return Response(self.get_serializer(instance).data)
        
        logger.info(f'创建公司信息: user_id={request.user.id}, data={request.data}')
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        """Update company info - only allowed if rejected or withdrawn"""
        import logging
        logger = logging.getLogger(__name__)
        
        instance = self.get_object()
        
        # Only owner or admin can update
        if instance.user != request.user and not request.user.is_staff:
            logger.warning(f'无权限修改公司信息: user_id={request.user.id}, company_info_id={instance.id}')
            return Response(
                {"error": "无权限修改"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Users can only update if rejected or withdrawn
        if not request.user.is_staff and instance.status == 'approved':
            logger.warning(f'已审核通过的信息不可修改: user_id={request.user.id}, company_info_id={instance.id}')
            return Response(
                {"error": "已审核通过的信息不可修改"},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not request.user.is_staff and instance.status == 'pending':
            logger.warning(f'审核中的信息不可修改: user_id={request.user.id}, company_info_id={instance.id}')
            return Response(
                {"error": "审核中不可修改，请先撤回"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # If user is updating rejected/withdrawn info, reset status to pending
        if not request.user.is_staff and instance.status in ['rejected', 'withdrawn']:
            logger.info(f'用户重新提交公司信息: user_id={request.user.id}, company_info_id={instance.id}')
            instance.status = 'pending'
            instance.approved_at = None
            instance.reject_reason = ''
        
        response = super().update(request, *args, **kwargs)
        
        # Save status change if it was modified
        if not request.user.is_staff and instance.status == 'pending':
            instance.save(update_fields=['status', 'approved_at', 'reject_reason', 'updated_at'])
        
        return response
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve company info and upgrade user to dealer"""
        from django.utils import timezone
        from decimal import Decimal
        from .credit_services import CreditAccountService
        
        company_info = self.get_object()
        
        if company_info.status != 'pending':
            return Response(
                {"error": "当前状态不可审核通过"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update company info status
        company_info.status = 'approved'
        company_info.approved_at = timezone.now()
        if company_info.reject_reason:
            company_info.reject_reason = ''
        company_info.save(update_fields=['status', 'approved_at', 'reject_reason', 'updated_at'])
        
        # Upgrade user to dealer
        user = company_info.user
        user.role = 'dealer'
        user.save()
        try:
            if not hasattr(user, 'credit_account'):
                CreditAccountService.create_credit_account(user, Decimal('0.00'), payment_term_days=30)
        except Exception:
            pass
        credit_account_data = None
        try:
            account = CreditAccount.objects.get(user=user)
            credit_account_data = CreditAccountSerializer(account).data
        except CreditAccount.DoesNotExist:
            credit_account_data = None
        
        return Response({
            "message": "审核通过，用户已升级为经销商",
            "company_info": CompanyInfoSerializer(company_info).data,
            "user": UserSerializer(user).data,
            "credit_account": credit_account_data
        })
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject company info submission"""
        company_info = self.get_object()
        
        if company_info.status != 'pending':
            return Response(
                {"error": "当前状态不可拒绝"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        reason = str(request.data.get('reason') or request.data.get('reject_reason') or '').strip()
        company_info.status = 'rejected'
        company_info.approved_at = None
        company_info.reject_reason = reason
        company_info.save(update_fields=['status', 'approved_at', 'reject_reason', 'updated_at'])
        
        return Response({
            "message": "已拒绝该公司信息",
            "company_info": CompanyInfoSerializer(company_info).data
        })

    @action(detail=True, methods=['post'])
    def withdraw(self, request, pk=None):
        """Withdraw pending company info submission"""
        company_info = self.get_object()

        if company_info.user != request.user and not request.user.is_staff:
            return Response(
                {"error": "无权限撤回"},
                status=status.HTTP_403_FORBIDDEN
            )

        if company_info.status != 'pending':
            return Response(
                {"error": "当前状态不可撤回"},
                status=status.HTTP_400_BAD_REQUEST
            )

        company_info.status = 'withdrawn'
        company_info.approved_at = None
        company_info.reject_reason = ''
        company_info.save(update_fields=['status', 'approved_at', 'reject_reason', 'updated_at'])

        return Response({
            "message": "已撤回审核，可修改后重新提交",
            "company_info": CompanyInfoSerializer(company_info).data
        })

    @extend_schema(
        operation_id='company_info_destroy',
        description='删除公司信息：如被其他数据引用时，返回提示信息',
    )
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            from django.db.models.deletion import ProtectedError
            if isinstance(e, ProtectedError):
                return Response(
                    {
                        'error': '无法删除公司信息',
                        'message': '该公司信息被其他数据引用，无法删除',
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            raise


@extend_schema(tags=['Credit Account'])
class CreditAccountViewSet(viewsets.ModelViewSet):
    """
    信用账户管理
    
    Admin endpoints:
    - GET list: 查看所有经销商信用账户
    - POST: 为经销商创建信用账户
    - PATCH: 更新信用额度和账期
    - GET detail: 查看信用账户详情
    
    Dealer endpoints:
    - GET my_account: 查看自己的信用账户
    """
    queryset = CreditAccount.objects.all().select_related('user', 'user__company_info').order_by('-created_at')
    serializer_class = CreditAccountSerializer
    
    def get_permissions(self):
        if self.action in ['my_account']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdmin()]
    
    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return CreditAccount.objects.none()
        
        if user.is_staff:
            # Admin can see all
            qs = super().get_queryset()
            # Filter by active status
            is_active = self.request.query_params.get('is_active')
            if is_active is not None:
                parsed = to_bool(is_active)
                if parsed is not None:
                    qs = qs.filter(is_active=parsed)
            return qs
        else:
            # Dealers can only see their own
            return CreditAccount.objects.filter(user=user)
    
    @action(detail=False, methods=['get'])
    def my_account(self, request):
        """经销商查看自己的信用账户"""
        try:
            account = CreditAccount.objects.get(user=request.user)
            serializer = self.get_serializer(account)
            return Response(serializer.data)
        except CreditAccount.DoesNotExist:
            return Response(
                {"error": "您还没有信用账户"},
                status=status.HTTP_404_NOT_FOUND
            )

    @extend_schema(
        operation_id='credit_accounts_destroy',
        description='删除信用账户：存在未结清欠款或关联账单/交易时，阻止删除并返回提示信息',
    )
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        statements_count = instance.statements.count()
        transactions_count = instance.transactions.count()
        has_debt = instance.outstanding_debt > 0
        if has_debt or statements_count > 0 or transactions_count > 0:
            return Response(
                {
                    'error': '无法删除信用账户',
                    'message': (
                        f"该账户{'存在未结清欠款，' if has_debt else ''}有 {statements_count} 条对账单、{transactions_count} 条账务交易，无法删除"
                    ).strip(),
                    'outstanding_debt': float(instance.outstanding_debt),
                    'statements_count': statements_count,
                    'transactions_count': transactions_count,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            from django.db.models.deletion import ProtectedError
            if isinstance(e, ProtectedError):
                return Response(
                    {
                        'error': '无法删除信用账户',
                        'message': '该账户被其他数据引用，无法删除',
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            raise

@extend_schema(tags=['Account Statement'])
class AccountStatementViewSet(viewsets.ModelViewSet):
    """
    账务对账单管理
    
    Admin endpoints:
    - GET list: 查看所有对账单
    - POST: 创建对账单
    - GET detail: 查看对账单详情（包含交易记录）
    - confirm: 确认对账单
    - settle: 结清对账单
    - export: 导出对账单为Excel
    
    Dealer endpoints:
    - GET my_statements: 查看自己的对账单列表
    - GET detail: 查看对账单详情
    """
    queryset = AccountStatement.objects.all().select_related(
        'credit_account',
        'credit_account__user',
        'credit_account__user__company_info'
    ).order_by('-period_end', '-created_at')
    
    def get_permissions(self):
        if self.action in ['my_statements', 'retrieve', 'confirm']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdmin()]
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AccountStatementDetailSerializer
        return AccountStatementSerializer
    
    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return AccountStatement.objects.none()
        
        if user.is_staff:
            # Admin can see all
            qs = super().get_queryset()

            search = self.request.query_params.get('search')
            if search:
                qs = qs.filter(
                    Q(credit_account__user__username__icontains=search) |
                    Q(credit_account__user__company_info__company_name__icontains=search)
                )

            # Filter by status
            status_filter = self.request.query_params.get('status')
            if status_filter:
                qs = qs.filter(status=status_filter)
            
            # Filter by credit account
            credit_account_id = self.request.query_params.get('credit_account')
            if credit_account_id:
                qs = qs.filter(credit_account_id=credit_account_id)
            
            # Filter by date range
            start_date = self.request.query_params.get('start_date')
            end_date = self.request.query_params.get('end_date')
            if start_date:
                qs = qs.filter(period_start__gte=start_date)
            if end_date:
                qs = qs.filter(period_end__lte=end_date)
            
            return qs
        else:
            # Dealers can only see their own
            try:
                credit_account = CreditAccount.objects.get(user=user)
                return AccountStatement.objects.filter(credit_account=credit_account)
            except CreditAccount.DoesNotExist:
                return AccountStatement.objects.none()

    @action(detail=False, methods=['get'], url_path='export')
    def export_list(self, request):
        qs = self.filter_queryset(self.get_queryset())
        headers = [
            'ID',
            '经销商',
            '公司名称',
            '账期开始',
            '账期结束',
            '期末未付',
            '逾期金额',
            '状态',
            '创建时间',
        ]
        rows = []
        for statement in qs:
            rows.append([
                statement.id,
                statement.credit_account.user.username if statement.credit_account else '',
                getattr(getattr(statement.credit_account.user, 'company_info', None), 'company_name', ''),
                statement.period_start,
                statement.period_end,
                statement.period_end_balance,
                statement.overdue_amount,
                statement.get_status_display(),
                statement.created_at,
            ])
        filename = f"statements_export_{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return build_excel_response(filename, headers, rows, title="对账单导出")

    def create(self, request, *args, **kwargs):
        """管理员创建对账单：根据账期自动汇总交易生成对账单"""
        from .credit_services import AccountStatementService
        from datetime import date
        try:
            credit_account_id = request.data.get('credit_account')
            period_start = request.data.get('period_start')
            period_end = request.data.get('period_end')
            if not credit_account_id or not period_start or not period_end:
                return Response({"error": "credit_account、period_start、period_end 为必填"}, status=status.HTTP_400_BAD_REQUEST)
            # Parse dates
            try:
                ps = date.fromisoformat(str(period_start))
                pe = date.fromisoformat(str(period_end))
            except Exception:
                return Response({"error": "账期日期格式错误，需使用YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)
            if ps > pe:
                return Response({"error": "账期开始不能晚于账期结束"}, status=status.HTTP_400_BAD_REQUEST)
            # Load credit account
            try:
                credit_account = CreditAccount.objects.select_related('user').get(id=credit_account_id)
            except CreditAccount.DoesNotExist:
                return Response({"error": "信用账户不存在"}, status=status.HTTP_404_NOT_FOUND)
            # Only allow for dealers
            if credit_account.user.role != 'dealer':
                return Response({"error": "仅支持为经销商创建对账单"}, status=status.HTTP_400_BAD_REQUEST)
            # Prevent duplicate statement for exact period
            existing = AccountStatement.objects.filter(
                credit_account=credit_account,
                period_start=ps,
                period_end=pe,
            ).exists()
            if existing:
                return Response({"error": "该账期的对账单已存在"}, status=status.HTTP_400_BAD_REQUEST)
            # Generate statement via service
            statement = AccountStatementService.generate_statement(credit_account, ps, pe)
            try:
                from users.services import create_notification
                create_notification(
                    credit_account.user,
                    title='新的对账单已生成',
                    content=f'{ps} 至 {pe} 的对账单已生成，期末未付 ¥{statement.period_end_balance}',
                    ntype='statement',
                    metadata={
                        'statement_id': statement.id,
                        'status': statement.status,
                        'period_start': str(ps),
                        'period_end': str(pe),
                        'amount': str(statement.period_end_balance),
                        'page': f'pages/statement-detail/index?id={statement.id}',
                        'subscription_data': {
                            'thing1': {'value': f'{ps} 至 {pe}'[:20]},
                            'time2': {'value': timezone.localtime(statement.created_at).strftime('%Y-%m-%d %H:%M') if statement.created_at else ''},
                            'thing3': {'value': f'未付 ¥{statement.period_end_balance}'[:20]},
                        },
                    }
                )
            except Exception:
                pass
            serializer = self.get_serializer(statement)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def my_statements(self, request):
        """经销商查看自己的对账单列表"""
        try:
            credit_account = CreditAccount.objects.get(user=request.user)
            statements = AccountStatement.objects.filter(
                credit_account=credit_account
            ).order_by('-period_end')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if start_date:
                statements = statements.filter(period_start__gte=start_date)
            if end_date:
                statements = statements.filter(period_end__lte=end_date)
            
            # Apply pagination
            page = self.paginate_queryset(statements)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(statements, many=True)
            return Response(serializer.data)
        except CreditAccount.DoesNotExist:
            return Response(
                {"error": "您还没有信用账户"},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """确认对账单"""
        from django.utils import timezone
        
        statement = self.get_object()
        
        if statement.status != 'draft':
            return Response(
                {"error": "只能确认草稿状态的对账单"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        statement.status = 'confirmed'
        statement.confirmed_at = timezone.now()
        statement.save()

        try:
            from users.services import create_notification
            create_notification(
                statement.credit_account.user,
                title='对账单已确认',
                content=f'{statement.period_start} 至 {statement.period_end} 的对账单已确认',
                ntype='statement',
                metadata={
                    'statement_id': statement.id,
                    'status': statement.status,
                    'period_start': str(statement.period_start),
                    'period_end': str(statement.period_end),
                    'amount': str(statement.period_end_balance),
                    'page': f'pages/statement-detail/index?id={statement.id}',
                    'subscription_data': {
                        'thing1': {'value': f'{statement.period_start} 至 {statement.period_end}'[:20]},
                        'time2': {'value': timezone.localtime(statement.confirmed_at).strftime('%Y-%m-%d %H:%M') if statement.confirmed_at else ''},
                        'thing3': {'value': '对账单已确认'},
                    },
                }
            )
        except Exception:
            pass
        
        return Response({
            "message": "对账单已确认",
            "statement": AccountStatementSerializer(statement).data
        })
    
    @action(detail=True, methods=['post'])
    def settle(self, request, pk=None):
        """结清对账单"""
        from django.utils import timezone
        
        statement = self.get_object()
        
        if statement.status == 'settled':
            return Response(
                {"error": "对账单已结清"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        statement.status = 'settled'
        statement.settled_at = timezone.now()
        statement.save()
        
        # Update credit account outstanding debt
        credit_account = statement.credit_account
        credit_account.outstanding_debt -= statement.period_end_balance
        if credit_account.outstanding_debt < 0:
            credit_account.outstanding_debt = 0
        credit_account.save()
        
        # Create payment transaction
        AccountTransaction.objects.create(
            credit_account=credit_account,
            statement=statement,
            transaction_type='payment',
            amount=statement.period_end_balance,
            balance_after=credit_account.outstanding_debt,
            payment_status='paid',
            paid_date=timezone.now().date(),
            description=f'对账单结算: {statement.period_start} 至 {statement.period_end}'
        )

        try:
            from users.services import create_notification
            create_notification(
                credit_account.user,
                title='对账单已结清',
                content=f'{statement.period_start} 至 {statement.period_end} 的对账单已结清',
                ntype='statement',
                metadata={
                    'statement_id': statement.id,
                    'status': statement.status,
                    'period_start': str(statement.period_start),
                    'period_end': str(statement.period_end),
                    'amount': str(statement.period_end_balance),
                    'page': f'pages/statement-detail/index?id={statement.id}',
                    'subscription_data': {
                        'thing1': {'value': f'{statement.period_start} 至 {statement.period_end}'[:20]},
                        'time2': {'value': timezone.localtime(statement.settled_at).strftime('%Y-%m-%d %H:%M') if statement.settled_at else ''},
                        'thing3': {'value': '对账单已结清'},
                    },
                }
            )
        except Exception:
            pass
        
        return Response({
            "message": "对账单已结清",
            "statement": AccountStatementSerializer(statement).data
        })
    
    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """导出对账单为Excel"""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        statement = self.get_object()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "对账单"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"账务对账单 - {statement.credit_account.user.company_info.company_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Statement info
        ws['A3'] = "账期："
        ws['B3'] = f"{statement.period_start} 至 {statement.period_end}"
        ws['A4'] = "状态："
        ws['B4'] = statement.get_status_display()
        
        # Financial summary
        ws['A6'] = "财务汇总"
        ws['A6'].font = Font(bold=True)
        
        summary_data = [
            ["上期结余", statement.previous_balance],
            ["本期采购", statement.current_purchases],
            ["本期付款", statement.current_payments],
            ["本期退款", statement.current_refunds],
            ["期末未付", statement.period_end_balance],
            ["账期内应付", statement.due_within_term],
            ["账期内已付", statement.paid_within_term],
            ["往来余额（逾期）", statement.overdue_amount],
        ]
        
        for idx, (label, value) in enumerate(summary_data, start=7):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = float(value)
        
        # Transaction details
        ws[f'A{len(summary_data) + 9}'] = "交易明细"
        ws[f'A{len(summary_data) + 9}'].font = Font(bold=True)
        
        # Transaction headers
        headers = ["日期", "交易类型", "金额", "余额", "订单ID", "应付日期", "实付日期", "付款状态", "备注"]
        header_row = len(summary_data) + 10
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Transaction data
        transactions = statement.transactions.all().order_by('created_at')
        for row_idx, transaction in enumerate(transactions, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=transaction.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=transaction.get_transaction_type_display())
            ws.cell(row=row_idx, column=3, value=float(transaction.amount))
            ws.cell(row=row_idx, column=4, value=float(transaction.balance_after))
            ws.cell(row=row_idx, column=5, value=transaction.order_id or '')
            ws.cell(row=row_idx, column=6, value=transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '')
            ws.cell(row=row_idx, column=7, value=transaction.paid_date.strftime('%Y-%m-%d') if transaction.paid_date else '')
            ws.cell(row=row_idx, column=8, value=transaction.get_payment_status_display())
            ws.cell(row=row_idx, column=9, value=transaction.description)
        
        # Adjust column widths
        for idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="statement_{statement.id}_{statement.period_start}_{statement.period_end}.xlsx"'
        wb.save(response)
        
        return response

    @extend_schema(
        operation_id='account_statements_destroy',
        description='删除对账单：如被其他数据引用时，返回提示信息',
    )
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            from django.db.models.deletion import ProtectedError
            if isinstance(e, ProtectedError):
                return Response(
                    {
                        'error': '无法删除对账单',
                        'message': '该对账单被其他数据引用，无法删除',
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            raise


@extend_schema(tags=['Account Transaction'])
class AccountTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    账务交易记录查询
    
    Admin endpoints:
    - GET list: 查看所有交易记录
    - GET detail: 查看交易详情
    
    Dealer endpoints:
    - GET my_transactions: 查看自己的交易记录
    """
    queryset = AccountTransaction.objects.all().select_related(
        'credit_account',
        'credit_account__user',
        'statement'
    ).order_by('-created_at')
    serializer_class = AccountTransactionSerializer
    
    def get_permissions(self):
        if self.action in ['my_transactions']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdmin()]
    
    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return AccountTransaction.objects.none()
        
        if user.is_staff:
            # Admin can see all
            qs = super().get_queryset()

            search = self.request.query_params.get('search')
            if search:
                qs = qs.filter(
                    Q(credit_account__user__username__icontains=search) |
                    Q(credit_account__user__company_info__company_name__icontains=search)
                )
            
            # Filter by credit account
            credit_account_id = self.request.query_params.get('credit_account')
            if credit_account_id:
                qs = qs.filter(credit_account_id=credit_account_id)
            
            # Filter by transaction type
            transaction_type = self.request.query_params.get('transaction_type')
            if transaction_type:
                qs = qs.filter(transaction_type=transaction_type)
            
            # Filter by payment status
            payment_status = self.request.query_params.get('payment_status')
            if payment_status:
                qs = qs.filter(payment_status=payment_status)
            
            # Filter by date range
            start_date = self.request.query_params.get('start_date')
            end_date = self.request.query_params.get('end_date')
            if start_date:
                qs = qs.filter(created_at__gte=start_date)
            if end_date:
                # Handle end_date inclusive for DateTimeField
                try:
                    from datetime import datetime, timedelta
                    ed = datetime.strptime(end_date, '%Y-%m-%d')
                    ed_end = ed + timedelta(days=1)
                    qs = qs.filter(created_at__lt=ed_end)
                except (ValueError, TypeError):
                    qs = qs.filter(created_at__lte=end_date)
            
            return qs
        else:
            # Dealers can only see their own
            try:
                credit_account = CreditAccount.objects.get(user=user)
                return AccountTransaction.objects.filter(credit_account=credit_account)
            except CreditAccount.DoesNotExist:
                return AccountTransaction.objects.none()

    @action(detail=False, methods=['get'])
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())
        headers = [
            'ID',
            '经销商',
            '公司名称',
            '交易类型',
            '金额',
            '变动后余额',
            '订单ID',
            '付款状态',
            '应付日期',
            '实付日期',
            '交易时间',
            '备注',
        ]
        rows = []
        for txn in qs:
            user = txn.credit_account.user if txn.credit_account else None
            company_name = getattr(getattr(user, 'company_info', None), 'company_name', '') if user else ''
            rows.append([
                txn.id,
                user.username if user else '',
                company_name,
                txn.get_transaction_type_display(),
                txn.amount,
                txn.balance_after,
                txn.order_id or '',
                txn.get_payment_status_display() if txn.transaction_type == 'purchase' else '-',
                txn.due_date,
                txn.paid_date,
                txn.created_at,
                txn.description,
            ])
        filename = f"transactions_export_{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return build_excel_response(filename, headers, rows, title="交易记录导出")
    
    @action(detail=False, methods=['get'])
    def my_transactions(self, request):
        """经销商查看自己的交易记录"""
        try:
            credit_account = CreditAccount.objects.get(user=request.user)
            transactions = AccountTransaction.objects.filter(
                credit_account=credit_account
            ).order_by('-created_at')
            
            transaction_type = request.query_params.get('transaction_type')
            if transaction_type:
                transactions = transactions.filter(transaction_type=transaction_type)
            
            payment_status = request.query_params.get('payment_status')
            if payment_status:
                transactions = transactions.filter(payment_status=payment_status)
            
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if start_date:
                transactions = transactions.filter(created_at__gte=start_date)
            if end_date:
                # Handle end_date inclusive for DateTimeField
                try:
                    from datetime import datetime, timedelta
                    ed = datetime.strptime(end_date, '%Y-%m-%d')
                    ed_end = ed + timedelta(days=1)
                    transactions = transactions.filter(created_at__lt=ed_end)
                except (ValueError, TypeError):
                    transactions = transactions.filter(created_at__lte=end_date)
            
            # Apply pagination
            page = self.paginate_queryset(transactions)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(transactions, many=True)
            return Response(serializer.data)
        except CreditAccount.DoesNotExist:
            return Response(
                {"error": "您还没有信用账户"},
                status=status.HTTP_404_NOT_FOUND
            )
